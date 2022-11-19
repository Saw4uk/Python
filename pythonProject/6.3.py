import csv
import re
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl import Workbook
from jinja2 import Environment, FileSystemLoader
import pdfkit
import excel2img

clear_html = re.compile('<.*?>')
format_money = re.compile('\d+(\.\d{1,2})?')

#Коммент, подтверждающий, что файл из ветки develop
def split_skills(skills):
    return skills.split("\n")


class DataSet:

    def clear_file(self, file_massive):
        file_massive = self.clear_no_full_rows(file_massive)
        for i, row in enumerate(file_massive):
            file_massive[i] = self.clear_row(row)
        return file_massive

    @staticmethod
    def clear_row(row: list):
        for i, string in enumerate(row):
            if i == 2:
                continue
            row[i] = re.sub(clear_html, '', row[i])
            row[i] = " ".join(row[i].split())
        return row

    def clear_no_full_rows(self, file):
        return [row for row in file if self.is_full(row)]

    @staticmethod
    def is_full(row):
        if len(row) == 0:
            return False
        for string in row:
            if string == '':
                return False
        return True

    def CreateVacancy(self, list):
        return Vacancy(list, self.dic_keys)

    @staticmethod
    def read_file(name):
        with open(name, 'r', encoding='utf-8-sig') as f:
            file_massive = list(csv.reader(f))
        if len(file_massive) == 1:
            print("Нет данных")
            exit()
        if len(file_massive) == 0:
            print("Пустой файл")
            exit()
        return file_massive.pop(0), file_massive

    def __init__(self, file_name):
        self.file_name = file_name
        table_keys, file = self.read_file(file_name)
        self.dic_keys = {}
        for i in range(0, len(table_keys)):
            self.dic_keys[table_keys[i]] = i
        file = self.clear_file(file)
        self.vacancies_objects = list(map(self.CreateVacancy, file))


class Salary:
    @staticmethod
    def currency_to_rub():
        return {
            "AZN": 35.68,
            "BYR": 23.91,
            "EUR": 59.90,
            "GEL": 21.74,
            "KGS": 0.76,
            "KZT": 0.13,
            "RUR": 1,
            "UAH": 1.64,
            "USD": 60.66,
            "UZS": 0.0055,
        }

    @staticmethod
    def parse_money(money):
        x = re.sub(format_money, lambda x: "${:,.2f}"
                   .format(float(x.group())), money) \
            .replace(",", " ") \
            .replace("$", "") \
            .replace(".00", "")
        return x

    def get_formatted(self, dic_currency):
        return f'{Salary.parse_money(str(self.salary_from))} - {Salary.parse_money(str(self.salary_to))} ({dic_currency[self.salary_currency]}) ({"С вычетом налогов" if self.salary_gross is False else "Без вычета налогов"})'

    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.salary_middle = ((float(self.salary_from) * Salary.currency_to_rub()[salary_currency] + float(
            self.salary_to) * Salary.currency_to_rub()[salary_currency]) / 2.0)


class Vacancy:
    def __init__(self, vacancy: list, dic_keys):
        self.name = vacancy[dic_keys['name']]
        self.salary = Salary(vacancy[dic_keys['salary_from']], vacancy[dic_keys['salary_to']],
                             vacancy[dic_keys['salary_currency']])
        self.area_name = vacancy[dic_keys['area_name']]
        self.published_at = vacancy[dic_keys['published_at']]


class Statistics:

    def __init__(self, vacancy_name, data_set):
        self.total_count = len(data_set.vacancies_objects)

        def add_to_dic(amount_dic, salary_dic, vacancy, field):
            if field in salary_dic:
                salary_dic[field] += vacancy.salary.salary_middle
            else:
                salary_dic[field] = vacancy.salary.salary_middle
                amount_dic[field] = 1
                return
            amount_dic[field] += 1

        self.dic_towns_amount = {}
        self.dic_towns_salaries = {}

        self.dic_year_amount = {}
        self.dic_year_salaries = {}

        self.dic_vac_amount = {}
        self.dic_vac_salaries = {}

        for vacancy in data_set.vacancies_objects:
            vacancy_year = int(vacancy.published_at[0:4])
            if vacancy_name in vacancy.name:
                add_to_dic(self.dic_vac_amount, self.dic_vac_salaries, vacancy, vacancy_year)
            add_to_dic(self.dic_towns_amount, self.dic_towns_salaries, vacancy, vacancy.area_name)
            add_to_dic(self.dic_year_amount, self.dic_year_salaries, vacancy, vacancy_year)
            if len(self.dic_vac_amount) == 0:
                self.dic_vac_salaries[vacancy_year] = 0
                self.dic_vac_amount[vacancy_year] = 0
        result_dict = {}
        for town in self.dic_towns_amount.keys():
            if self.dic_towns_amount[town] / self.total_count >= 0.01:
                result_dict[town] = self.dic_towns_amount[town]
            else:
                self.dic_towns_salaries.pop(town)
        self.dic_towns_amount = result_dict

    def get_format_dic(self, dic, amdic):
        resultdic = {}
        for year in dic.keys():
            if amdic[year] == 0:
                resultdic[year] = 0
                return resultdic
            resultdic[year] = int(dic[year] / amdic[year])
        return resultdic

    def get_ten_len(self, dic):
        result = {}
        for x in dic.keys():
            result[x] = dic[x]
            if len(result) == 10:
                return result
        return result

    def get_parts(self):
        dic_parts = {}
        for town in self.dic_towns_amount.keys():
            to_add = self.dic_towns_amount[town] / self.total_count
            if to_add > 0.01:
                dic_parts[town] = round(to_add, 4)
            else:
                continue
        return self.get_ten_len(self.sort_dic(dic_parts))

    def sort_dic(self, dic: dict):
        return dict(sorted(dic.items(), key=lambda item: float(item[1]), reverse=True))

    def print_formatted(self):
        self.final_year_salar = self.get_format_dic(self.dic_year_salaries, self.dic_year_amount)
        self.final_year_amount = self.dic_year_amount
        print(f"Динамика уровня зарплат по годам: {self.final_year_salar}")
        print(f"Динамика количества вакансий по годам: {self.final_year_amount}")
        self.final_vac_salar = self.get_format_dic(self.dic_vac_salaries, self.dic_vac_amount)
        self.final_vac_amount = self.dic_vac_amount
        print(
            f"Динамика уровня зарплат по годам для выбранной профессии: {self.final_vac_salar}")
        print(f"Динамика количества вакансий по годам для выбранной профессии: {self.final_vac_amount}")
        self.final_town_salar = self.get_ten_len(
            self.sort_dic(self.get_format_dic(self.dic_towns_salaries, self.dic_towns_amount)))
        self.final_town_amount = self.get_ten_len((self.get_parts()))
        print(
            f"Уровень зарплат по городам (в порядке убывания): {self.final_town_salar}")
        print(f"Доля вакансий по городам (в порядке убывания): {self.final_town_amount}")


class DiagrammCreator:
    @staticmethod
    def create_collumns_diagramm(years, vacancy_salaries_array, middle_salaries_array, vacancy_name, table_name, ax):
        x = np.arange(len(years))  # the label locations
        width = 0.3  # the width of the bars

        ax.bar(x - width / 2, middle_salaries_array, width, label='Средняя з/п')
        ax.bar(x + width / 2, vacancy_salaries_array, width, label=f'з/п {vacancy_name}')

        ax.grid(axis="y")
        ax.set_title(table_name)
        ax.set_xticks(x, years, rotation=90, fontsize=8)
        ax.legend(fontsize=8)

    @staticmethod
    def create_horizontal_diagramm(towns_array, towns_array_values, table_name, ax):
        y_pos = np.arange(len(towns_array))
        ax.grid(axis="y")
        ax.barh(y_pos, towns_array_values, align='center')
        ax.set_yticks(y_pos, labels=towns_array, fontsize=6)
        ax.invert_yaxis()  # labels read top-to-bottom
        ax.set_title(table_name)

    @staticmethod
    def create_circle_diagramm(sizes, labels, table_name, ax):
        ax.set_title(table_name)
        patches, texts = ax.pie([1 - sum(sizes)] + sizes, labels=["Другие"] + labels)
        for tick in texts:
            tick.set_fontsize(6)


class PDFCreator:
    @staticmethod
    def PDF_func(vac_name):
        options = {'enable-local-file-access': None}
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Users\Saw4uk\Downloads\wkhtmltox\bin\wkhtmltopdf.exe')

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template_html.html")

        pdf_template = template.render({'name': vac_name})

        pdfkit.from_string(pdf_template, 'out.pdf', configuration=config, options=options)


class StartProgramm:
    def __init__(self):
        file_name = "vacancies_by_year.csv"  # ("Введите название файла: ")
        vacancy_name = "Аналитик"  # input("Введите название профессии: ")
        data_set = DataSet(file_name)
        stat = Statistics(vacancy_name, data_set)
        stat.print_formatted()
        Report.generate_files(stat, vacancy_name)


class Report:
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    @staticmethod
    def make_main_row(coord, str, sheet):
        sheet[coord + '1'] = str
        sheet[coord + '1'].border = Report.thin_border
        sheet[coord + '1'].font = Font(
            name='Stalk',
            size=10,
            bold=True,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')

    @staticmethod
    def format_sheet_collumns_width(sheet):
        def as_text(value):
            if value is None:
                return ""
            return str(value)

        for column_cells in sheet.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 5

    @staticmethod
    def write_two_collumns_to_sheet(firstcoord, first_str, secondcoord, second_str, sheet, dict, shouldPercent):
        def to_percent_format(number):
            return str(round(number * 100, 2)).replace('.', ',') + "%"

        Report.make_main_row(firstcoord, first_str, sheet)
        Report.make_main_row(secondcoord, second_str, sheet)
        for row, (key, string) in enumerate(dict.items(), start=2):
            sheet[f"{firstcoord}{row}"] = key
            sheet[f"{firstcoord}{row}"].border = Report.thin_border

            sheet[f"{secondcoord}{row}"].border = Report.thin_border
            if shouldPercent:
                sheet[f"{secondcoord}{row}"] = to_percent_format(float(string))
            else:
                sheet[f"{secondcoord}{row}"] = string
        Report.format_sheet_collumns_width(sheet)

    @staticmethod
    def write_first_sheet(sheet, year_sal, year_amount, vac_sal, vac_amount, vac_name):
        def fill_collumn(dic, column):
            for x in range(0, len(year_sal)):
                sheet[column + str(x + 2)] = dic[sheet["A" + str(x + 2)].internal_value]
                sheet[column + str(x + 2)].border = Report.thin_border

        Report.make_main_row("A", "Год", sheet)
        Report.make_main_row("B", "Средняя зарплата", sheet)
        Report.make_main_row("C", f"Средняя зарплата - {vac_name}", sheet)
        Report.make_main_row("D", "Количество вакансий", sheet)
        Report.make_main_row("E", f"Количество вакансий - {vac_name}", sheet)
        for row, key in enumerate(year_sal.keys(), start=2):
            sheet[f"A{row}"] = key
            sheet[f"A{row}"].border = Report.thin_border
        fill_collumn(year_sal, "B")
        fill_collumn(vac_sal, "C")
        fill_collumn(year_amount, "D")
        fill_collumn(vac_sal, "E")
        Report.format_sheet_collumns_width(sheet)

    @staticmethod
    def generate_files(statistics: Statistics, vac_name):
        wb = Workbook()
        ##wb.remove(wb['Sheet'])
        year_statistics_sheet = wb.create_sheet("Статистика по годам")
        Report.write_first_sheet(year_statistics_sheet, statistics.final_year_amount, statistics.final_year_amount,
                                 statistics.final_vac_salar, statistics.final_vac_amount, vac_name)
        town_statistics_sheet = wb.create_sheet("Статистика по городам")
        Report.write_two_collumns_to_sheet("A", "Город", "B", "Уровень зарплат", town_statistics_sheet,
                                           statistics.final_town_salar, False)
        Report.write_two_collumns_to_sheet("D", "Город", "E", "Доля вакансий", town_statistics_sheet,
                                           statistics.final_town_amount, True)
        wb.save('report.xlsx')
        excel2img.export_img("report.xlsx", "table1.png", "Статистика по годам", None)
        excel2img.export_img("report.xlsx", "table2.png", "Статистика по городам", None)
        fig, ax = plt.subplots(2, 2)

        DiagrammCreator.create_collumns_diagramm(statistics.final_year_salar.keys(),
                                                 statistics.final_vac_salar.values(),
                                                 statistics.final_year_salar.values(), vac_name,
                                                 "Уровень зарплат по городам", ax[0, 0])
        DiagrammCreator.create_collumns_diagramm(statistics.final_year_salar.keys(),
                                                 statistics.final_vac_amount.values(),
                                                 statistics.final_year_amount.values(), vac_name,
                                                 "Количество выкансий по годам", ax[0, 1])
        DiagrammCreator.create_horizontal_diagramm(
            list(map(lambda x: x.replace(' ', '\n').replace('-', '\n'), statistics.final_town_amount.keys())),
            statistics.get_ten_len(statistics.sort_dic(statistics.dic_towns_amount)).values(),
            "Количество выкансий по годам", ax[1, 0])
        DiagrammCreator.create_circle_diagramm(list(statistics.final_town_amount.values()),
                                               list(statistics.final_town_amount.keys()), "Доля вакансий по городам",
                                               ax[1, 1])

        fig.tight_layout()

        plt.savefig("output.jpg")

        PDFCreator.PDF_func(vac_name)


StartProgramm()
